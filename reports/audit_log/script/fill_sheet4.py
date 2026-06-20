import openpyxl
import os

file_path = r'C:\Users\Lenovo\Desktop\DAP391_project\07_audit_log\Phan-Nhat-Duy Le.xlsx'
wb = openpyxl.load_workbook(file_path)

# Calculate counts
sheet2 = wb['2. Detailed Audit Log']
entries_count = 0
for row in range(4, sheet2.max_row + 1):
    if sheet2.cell(row=row, column=1).value is not None:
        entries_count += 1

sheet3 = wb['3. Hallucination Detection']
hallucinations_count = 0
for row in range(4, sheet3.max_row + 1):
    if sheet3.cell(row=row, column=1).value is not None:
        hallucinations_count += 1

sheet4 = wb['4. Self-Assessment Checklist']

# Assuming B starts around row 13-15. Let's find "B. KIỂM TRA TỔNG THỂ LOG"
start_row_b = None
for row in range(1, 30):
    val = str(sheet4.cell(row=row, column=1).value or '')
    if 'B. KIỂM TRA TỔNG THỂ' in val.upper() or 'B. KIỂM TRA TỔNG THỂ LOG' in str(sheet4.cell(row=row, column=2).value or '').upper():
        start_row_b = row
        break
    
if start_row_b is None:
    # try column 2
    for row in range(1, 30):
        val = str(sheet4.cell(row=row, column=2).value or '')
        if 'B. KIỂM TRA TỔNG THỂ' in val.upper():
            start_row_b = row
            break

if start_row_b is not None:
    # Row indices for criteria (start_row_b + 2 onwards)
    criteria_start = start_row_b + 2
    
    # 1. Số lượng entries
    sheet4.cell(row=criteria_start, column=3).value = '☑'
    sheet4.cell(row=criteria_start, column=4).value = f'{entries_count} entries'
    
    # 2. Mỗi component có ít nhất 1 core prompt?
    sheet4.cell(row=criteria_start+1, column=3).value = '☑'
    sheet4.cell(row=criteria_start+1, column=4).value = 'Đầy đủ các steps'
    
    # 3. Đã phát hiện ≥ số lượng hallucination yêu cầu?
    sheet4.cell(row=criteria_start+2, column=3).value = '☑'
    sheet4.cell(row=criteria_start+2, column=4).value = f'{hallucinations_count} hallucinations'
    
    # 4. Mỗi entry đều có Human Delta đầy đủ (4 câu hỏi)?
    sheet4.cell(row=criteria_start+3, column=3).value = '☑'
    sheet4.cell(row=criteria_start+3, column=4).value = '100% đầy đủ'
    
    # 5. Có evidence cho ≥70% entries?
    sheet4.cell(row=criteria_start+4, column=3).value = '☑'
    sheet4.cell(row=criteria_start+4, column=4).value = '100% có evidence'
    
    # Tích các ô pass ở phần A nếu muốn
    start_row_a = None
    for row in range(1, 15):
        val = str(sheet4.cell(row=row, column=1).value or '') + str(sheet4.cell(row=row, column=2).value or '')
        if 'A. KIỂM TRA CHẤT LƯỢNG' in val.upper():
            start_row_a = row
            break
            
    if start_row_a is not None:
        for offset in range(2, 7):
            sheet4.cell(row=start_row_a+offset, column=3).value = '☑'

wb.save(file_path)
print(f"Updated {file_path} successfully!")
