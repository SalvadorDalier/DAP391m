import openpyxl

file_path = r"C:\Users\Lenovo\Desktop\DAP391_project\07_audit_log\Audit_log_Restructured.xlsx"
wb = openpyxl.load_workbook(file_path)
sheet = wb["2. Detailed Audit Log"]

for row in range(4, 8):
    prompt = sheet.cell(row=row, column=4).value # Column D might be Prompt? Let's check columns first
    print(f"Row {row}:")
    for col in range(1, 10):
        print(f"  Col {col}: {sheet.cell(row=row, column=col).value}")

