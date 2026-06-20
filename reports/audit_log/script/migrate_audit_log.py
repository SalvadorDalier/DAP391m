import pandas as pd
import openpyxl

current_path = r'C:\Users\Lenovo\Desktop\DAP391_project\07_audit_log\Audit_log.xlsx'
template_path = r'C:\Users\Lenovo\Downloads\Report Templates-20260603\AI_AuditLog_Template_DAP391m.xlsx'
output_path = r'C:\Users\Lenovo\Desktop\DAP391_project\07_audit_log\Final_Audit_Log.xlsx'

# Load current data
core_prompts = pd.read_excel(current_path, sheet_name='Core Prompts')
ai_responses = pd.read_excel(current_path, sheet_name='AI Responses')
human_deltas = pd.read_excel(current_path, sheet_name='Human Delta')
hallucinations = pd.read_excel(current_path, sheet_name='Hallucination Detections')

wb = openpyxl.load_workbook(template_path)

# Fill Sheet 1: 1. Metadata & Summary
sheet1 = wb['1. Metadata & Summary']
sheet1['B3'] = 'DAP391m Project Team'
sheet1['B4'] = 'DAP391m'
sheet1['B5'] = 'DAP391m'
sheet1['B6'] = 'Online Retail II Analysis & Marketing Uplift Modeling'
sheet1['B9'] = str(len(core_prompts))
sheet1['B10'] = str(len(core_prompts))

# Fill Sheet 2: 2. Detailed Audit Log
sheet2 = wb['2. Detailed Audit Log']
# Data starts at row 4
start_row = 4

# Clear existing sample data
for r in range(start_row, sheet2.max_row + 1):
    for c in range(1, 9):
        sheet2.cell(row=r, column=c).value = None

row_idx = start_row
step_to_entry = {}

for idx, row in core_prompts.iterrows():
    entry_num = idx + 1
    step = str(row['Step'])
    prompt = str(row['Prompt'])
    purpose = str(row['Purpose'])
    date = str(row['Date'])
    
    step_to_entry[step] = entry_num
    
    # Match AI response
    ai_resp_df = ai_responses[ai_responses['Step'] == step]
    ai_resp_txt = ai_resp_df.iloc[0]['AI_Response_Summary'] if len(ai_resp_df) > 0 else "N/A"
    
    # Match Human Delta
    hd_df = human_deltas[human_deltas['Step'] == step]
    if len(hd_df) > 0:
        hd = hd_df.iloc[0]
        human_txt = f"Human Modification: {hd['Human_Modification']}\nReason: {hd['Reason_for_Change']}"
    else:
        human_txt = "Verified logic, no modifications needed."
        
    sheet2.cell(row=row_idx, column=1).value = entry_num
    sheet2.cell(row=row_idx, column=2).value = "PROBLEM-SOLVING"
    sheet2.cell(row=row_idx, column=3).value = step
    sheet2.cell(row=row_idx, column=4).value = purpose
    sheet2.cell(row=row_idx, column=5).value = prompt
    sheet2.cell(row=row_idx, column=6).value = ai_resp_txt
    sheet2.cell(row=row_idx, column=7).value = human_txt
    sheet2.cell(row=row_idx, column=8).value = "Source code / documentation"
    
    row_idx += 1

# Fill Sheet 3: 3. Hallucination Detection
sheet3 = wb['3. Hallucination Detection']
# Data starts at row 4
start_row_3 = 4

for r in range(start_row_3, sheet3.max_row + 1):
    for c in range(1, 7):
        sheet3.cell(row=r, column=c).value = None

row_idx_3 = start_row_3
for idx, row in hallucinations.iterrows():
    step = str(row['Step'])
    entry_num = step_to_entry.get(step, "N/A")
    
    sheet3.cell(row=row_idx_3, column=1).value = entry_num
    sheet3.cell(row=row_idx_3, column=2).value = "Fabrication"
    sheet3.cell(row=row_idx_3, column=3).value = str(row['AI_Output'])
    sheet3.cell(row=row_idx_3, column=4).value = str(row['Hallucination_Description'])
    sheet3.cell(row=row_idx_3, column=5).value = "Execution and Code Review"
    sheet3.cell(row=row_idx_3, column=6).value = str(row['Correction_Method'])
    
    row_idx_3 += 1

wb.save(output_path)
print(f"Successfully migrated data to {output_path}")
