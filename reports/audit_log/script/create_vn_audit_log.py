import openpyxl

template_path = r'C:\Users\Lenovo\Downloads\Report Templates-20260603\AI_AuditLog_Template_DAP391m.xlsx'
output_path = r'C:\Users\Lenovo\Desktop\DAP391_project\07_audit_log\Audit_log_Restructured_VN.xlsx'

wb = openpyxl.load_workbook(template_path)

# Sheet 1: Metadata
sheet1 = wb['1. Metadata & Summary']
sheet1['B3'] = 'Nhóm dự án DAP391m'
sheet1['B4'] = 'DAP391m'
sheet1['B5'] = 'DAP391m'
sheet1['B6'] = 'Phân tích Online Retail II & Marketing Uplift Modeling'
sheet1['B9'] = '4'
sheet1['B10'] = '4'

# Sheet 2: Detailed Audit Log
sheet2 = wb['2. Detailed Audit Log']
start_row = 4
for r in range(start_row, sheet2.max_row + 1):
    for c in range(1, 9):
        sheet2.cell(row=r, column=c).value = None

data_s2 = [
    (1, "PROBLEM-SOLVING", "Step 1 & 2 - Data Import & Cleaning", 
     "Mục đích: Thiết lập pipeline ETL tự động để làm sạch dữ liệu và dễ dàng đưa vào database.", 
     "Viết script Python để tiền xử lý tập dữ liệu Online Retail II: xử lý CustomerID bị thiếu, giảm ảnh hưởng của outlier trong Quantity và UnitPrice bằng cách giới hạn theo percentile, sau đó đưa dữ liệu đã xử lý vào SQL Server.", 
     "Cung cấp script import_to_sql.py sử dụng pandas và pyodbc để tiền xử lý file Excel thô (lọc bỏ bản ghi thiếu CustomerID và giới hạn outlier ở percentile 99) trước khi chèn vào bảng core.Customer và core.Transaction.", 
     "Tư duy phản biện: Logic xử lý dữ liệu của AI hợp lý và hiệu quả.\nBối cảnh: Phù hợp với cấu trúc database SQL Server của dự án.\nTổng hợp sáng tạo: Áp dụng code trực tiếp không cần sửa đổi nhiều.\nQuyết định: Chấp nhận hoàn toàn giải pháp ETL của AI.", 
     "Source code import_to_sql.py"),
     
    (2, "PROBLEM-SOLVING", "Step 3 - Data Visualization", 
     "Mục đích: Tạo các biểu đồ trực quan để đưa vào báo cáo phân tích khám phá dữ liệu (EDA) và đánh giá mô hình.", 
     "Viết các script Python sử dụng matplotlib và seaborn để tạo biểu đồ histogram cho uplift, biểu đồ so sánh RFM, đường cong Qini và ma trận phân tích chi phí - lợi ích.", 
     "Tạo file visualization.py để vẽ phân phối RFM và tỷ lệ sử dụng coupon, giúp tích hợp dễ dàng vào báo cáo học thuật và Power BI dashboard.", 
     "Tư duy phản biện: Các biểu đồ AI đề xuất bám sát yêu cầu nghiệp vụ.\nBối cảnh: Rất cần thiết cho việc giải thích Uplift Model trong báo cáo.\nTổng hợp sáng tạo: Tích hợp trực tiếp vào module trực quan hóa.\nQuyết định: Giữ nguyên thiết kế biểu đồ của AI.", 
     "Source code visualization.py"),
     
    (3, "DECISION", "Step 5 - Predictive Modelling", 
     "Mục đích: Xây dựng và tối ưu hóa mô hình machine learning chính để nhắm mục tiêu khách hàng.", 
     "Triển khai và huấn luyện mô hình uplift T-Learner sử dụng XGBoost trong Python. Tính toán đường cong Qini và AUUC để lấy các chỉ số hiệu suất chính.", 
     "Cung cấp train_uplift_model.py. Do lỗi xung đột thư viện khi cài CausalML trên Windows, AI đã tạo một mô hình T-Learner thủ công bằng cách sử dụng hai mô hình xgboost.XGBClassifier riêng biệt cho nhóm control và treatment.", 
     "Tư duy phản biện: AI đề xuất chia dữ liệu ngẫu nhiên (shuffle=True) nhưng điều này gây rò rỉ dữ liệu thời gian (data leakage). AI cũng in text tiếng Việt gây lỗi Unicode.\nBối cảnh: Dự án cần dự đoán dựa trên lịch sử. Môi trường là Windows PowerShell.\nTổng hợp sáng tạo: Chỉnh sửa script để chia theo trình tự thời gian (80% đầu cho train). Thêm $env:PYTHONIOENCODING='utf-8' vào lệnh chạy.\nQuyết định: Thay đổi cách chia dữ liệu và biến môi trường để mô hình chuẩn xác và tránh lỗi runtime.", 
     "Lịch sử commit / split_train_test.py"),
     
    (4, "PROBLEM-SOLVING", "Step 6 - API & App Development", 
     "Mục đích: Xây dựng giao diện người dùng trực quan giúp các bên liên quan dễ dàng đưa ra quyết định dựa trên dữ liệu.", 
     "Phát triển một ứng dụng web tương tác bằng Streamlit bao gồm các thanh trượt mô phỏng (what-if), dự báo ROI động và giao diện tra cứu từng khách hàng.", 
     "Cung cấp streamlit_app.py, giúp tính toán Expected Profit Uplift sử dụng CATE cùng với các biến chi phí coupon, có thể điều chỉnh động qua thanh trượt.", 
     "Tư duy phản biện: Streamlit app của AI hoạt động tốt cho việc mô phỏng.\nBối cảnh: Giúp các stakeholder dễ dàng tương tác với kết quả mô hình Uplift.\nTổng hợp sáng tạo: Tích hợp model đã train vào giao diện web.\nQuyết định: Triển khai ứng dụng Streamlit theo hướng dẫn của AI.", 
     "Source code streamlit_app.py")
]

for idx, row in enumerate(data_s2):
    row_idx = start_row + idx
    for col_idx, val in enumerate(row):
        sheet2.cell(row=row_idx, column=col_idx+1).value = val


# Sheet 3: Hallucination
sheet3 = wb['3. Hallucination Detection']
start_row_3 = 4
for r in range(start_row_3, sheet3.max_row + 1):
    for c in range(1, 7):
        sheet3.cell(row=r, column=c).value = None

data_s3 = [
    (3, "Fabrication", "AI báo cáo các chỉ số trong Step5_Report.md: Baseline Accuracy = 0.9861, ROC AUC = 1.00, Uplift AUUC = 0.8830, và Random AUUC = -0.5940.", 
     "Các chỉ số này hoàn toàn do AI bịa đặt. Khi chạy thực tế train_uplift_model.py, baseline accuracy là 0.7092, Uplift AUUC là -9.5090 và Random AUUC là -10.0581.", 
     "Chạy code và kiểm tra kết quả", 
     "Chạy lại script huấn luyện thủ công trên tập test để lấy kết quả thực tế, sau đó sửa lại tất cả các chỉ số này trong báo cáo markdown."),
     
    (3, "Fabrication (API)", "Gọi lệnh import một class không tồn tại: from causalml.inference.meta import XGBClassifierTLEarner", 
     "Class XGBClassifierTLEarner không tồn tại trong thư viện CausalML, đây là một API hallucination.", 
     "Lỗi Runtime ImportError", 
     "Refactor lại code để sử dụng xgboost.XGBClassifier tiêu chuẩn cho mô hình Control và Treatment thay vì dùng CausalML.")
]

for idx, row in enumerate(data_s3):
    row_idx = start_row_3 + idx
    for col_idx, val in enumerate(row):
        sheet3.cell(row=row_idx, column=col_idx+1).value = val

wb.save(output_path)
print(f"Saved VN audit log to {output_path}")
